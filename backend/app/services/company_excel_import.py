"""Parse company bulk-import Excel (.xlsx) with flexible column headers."""
from __future__ import annotations

from typing import BinaryIO

from openpyxl import load_workbook

# Keep in sync with BULK_MAX_ITEMS in app.api.companies
_MAX_EXCEL_DATA_ROWS = 2000

# Normalized header (lowercase, spaces removed) → we match these sets
_NAME_HEADERS = frozenset({"name", "companyname", "company", "company_name"})
_KAM_HEADERS = frozenset(
    {
        "keyaccountmanager",
        "keyaccountmanagername",
        "kam",
        "kamname",
        "kamuser",
        "kamusername",
    }
)
_LOCATION_HEADERS = frozenset({"location", "city", "area"})
_COUNTRY_HEADERS = frozenset({"country", "county", "nation"})


def _norm_header(cell) -> str:
    if cell is None:
        return ""
    return "".join(str(cell).strip().lower().split())


def _find_column_index(header_row: tuple | list, candidates: frozenset[str]) -> int | None:
    for i, cell in enumerate(header_row):
        nh = _norm_header(cell)
        if nh in candidates:
            return i
    return None


def parse_company_excel(buffer: BinaryIO, filename: str) -> list[dict]:
    """
    Read first worksheet. First row = headers.

    Expected columns (flexible names):
    - Name / Company / Company name
    - KeyAccountManager / KAM / kam name → mapped to kamUserName for API
    - Location
    - Country or County

    Returns list of dicts: name, location, country, kamUserName (strings, may be empty where invalid).
    """
    if not filename or not str(filename).lower().endswith((".xlsx", ".xlsm")):
        raise ValueError("Upload an Excel file (.xlsx or .xlsm).")

    buffer.seek(0)
    wb = load_workbook(buffer, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            raise ValueError("Excel file has no rows.")

        hi_name = _find_column_index(header_row, _NAME_HEADERS)
        hi_kam = _find_column_index(header_row, _KAM_HEADERS)
        hi_loc = _find_column_index(header_row, _LOCATION_HEADERS)
        hi_cty = _find_column_index(header_row, _COUNTRY_HEADERS)

        missing = []
        if hi_name is None:
            missing.append("Name (or Company)")
        if hi_kam is None:
            missing.append("KeyAccountManager (or KAM)")
        if hi_loc is None:
            missing.append("Location")
        if hi_cty is None:
            missing.append("Country (or County)")
        if missing:
            raise ValueError(
                "Missing required column(s): " + ", ".join(missing) + ". "
                "First row must contain headers like Name, KeyAccountManager, Location, County/Country."
            )

        out: list[dict] = []
        for row in rows_iter:
            if row is None:
                continue
            def cell(i: int | None):
                if i is None or i >= len(row):
                    return ""
                v = row[i]
                if v is None:
                    return ""
                return str(v).strip()

            name = cell(hi_name)
            if not name:
                continue
            if len(out) >= _MAX_EXCEL_DATA_ROWS:
                raise ValueError(f"At most {_MAX_EXCEL_DATA_ROWS} data rows allowed.")

            out.append(
                {
                    "name": name,
                    "location": cell(hi_loc),
                    "country": cell(hi_cty),
                    "kamUserName": cell(hi_kam),
                }
            )
        return out
    finally:
        wb.close()
